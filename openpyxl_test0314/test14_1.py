from openpyxl import load_workbook
from openpyxl import workbook
from test0316_1 import Conf_test

class Excel_test:
    def __init__(self,ex,sh):
        wb=workbook.Workbook()
        wb.create_sheet(sh,index=0)
        wb.save(ex)
        self.data1='白日依山尽'
        self.data2="{'name':'邢麦苗','age':25}"
        self.data3='[0,2,3]'
        self.data4='黄河入海流'
        self.data5=25
        self.ex = ex
        self.sh = sh

    def read_excel(self,sheet_name):
        wb=load_workbook(self.ex)
        sheet=wb[self.sh]
        list1=[]
        # read_data=[]
        sheet_n = eval(sheet_name)
        for r in range(1, sheet.max_row + 1):
            list2 = []
            for c in range(1, sheet.max_column + 1):
                res = sheet.cell(r, c).value
                # print(res)
                list2.append(res)
            list1.append(list2)
        if sheet_name == 'all':
            read_data = list1[:]
        else:
            read_data = list1[0:sheet_n[-1] ]
        wb.close()
        return read_data

    def write_excel(self):
        wb=load_workbook(self.ex)
        sheet=wb[self.sh]
        sheet.cell(1,1).value=self.data1
        sheet.cell(1,3).value=self.data2
        sheet.cell(2,4).value=self.data3
        sheet.cell(2,5).value=self.data4
        sheet.cell(3,1).value=self.data5
        wb.save(self.ex)
        wb.close()

if __name__ == '__main__':


    excel='pyexcel_1.xlsx'
    sheet='xmm1'
    Ex = Excel_test(excel,sheet)
    Ex.write_excel()
    line = Conf_test('D:/PycharmProjects/2-24test/configparser_test0316/py_test0318.conf').get_steValue('excel_test','read')
    print(line)
    result=Ex.read_excel(line)
    print(result)


#问题：能否过滤列表中的中文？