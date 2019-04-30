from openpyxl import load_workbook
from common.requestHttp import RequestHttp

class Case:
    def __init__(self):
        '''測試用例類，每個測試用例，實際上就是它的一個實例'''
        self.id = None
        self.title = None
        self.url = None
        self.data = None
        self.method = None
        self.expectid = None
        self.actual = None
        self.result = None

class Do_Excel:
    def __init__(self,filename,sheet_name):
        # try:
        #     self.filename = filename
        #     self.wb = load_workbook(filename)
        #     self.sheet = self.wb[sheet_name]
        # except FileNotFoundError as e:
        #     raise e
        self.filename = filename
        self.sheet_name = sheet_name
        self.wb = load_workbook(filename)
        self.sheet = self.wb[sheet_name]



    def read_excel(self):
        self.wb = load_workbook(self.filename)
        self.sheet = self.wb[self.sheet_name]
        cases=[]
        # case = Case()    #注意：在這裡實例化只會取到最後一行數據
        for row in range(2,self.sheet.max_row+1):
            case = Case()
            case.id = self.sheet.cell(row, 1).value
            case.title = self.sheet.cell(row, 2).value
            case.url = self.sheet.cell(row, 3).value
            case.data = self.sheet.cell(row, 4).value
            case.method = self.sheet.cell(row, 5).value
            case.expectid = self.sheet.cell(row, 6).value
            case.sql = self.sheet.cell(row, 9).value
            cases.append(case)
        return cases

    def write_excel(self,row,actual,result):
        self.sheet.cell(row,7).value = actual
        self.sheet.cell(row,8).value = result
        self.wb.save(self.filename)
        self.wb.close()

if __name__ == '__main__':
    from cookies0409.UseExcel.common import contants
    try:
        excel = Do_Excel(contants.case_dir,'login')
        cases = excel.read_excel()

    except FileNotFoundError as e:
        print(e)
    else:
        print(cases)
        for case in cases:
            print(case.__dict__)   #此處是類對象的__dict__，主要包含init里定義的屬性
            print(type(case.data))
            resp = RequestHttp().request(case.method,case.url,case.data)
            if resp.text == case.expectid:
                excel.write_excel(case.id + 1,resp.text,'pass')
            else:
                excel.write_excel(case.id + 1, resp.text, 'false')



