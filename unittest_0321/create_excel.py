from openpyxl import workbook
from openpyxl import load_workbook

class Create_Excel:
    def __init__(self,sheetname,excelname):
        self.sheetname = sheetname
        self.excelname = excelname
        wb = workbook.Workbook()
        wb.create_sheet(self.sheetname,index=0)
        wb.save(self.excelname)
        self.op = load_workbook(self.excelname)
        self.sheet = self.op[self.sheetname]
        self.data1='{"mobilephone":"18688773467","pwd":"123456"}'
        self.data2='登录成功'
        self.data3='{"mobilephone":"18688773467","pwd":"1234567"}'
        self.data4="用户名或密码错误"
        self.data5='{"mobilephone":"18688773467","pwd":None}'
        self.data6="用户名或密码不能为空"
        self.data7='{"mobilephone":None,"pwd":"123456"}'
        self.data8="用户名或密码不能为空"
        self.data9 = 1
        self.data10 = 2
        self.data11= 3
        self.data12= 4

    def write_data(self):
        self.sheet.cell(1,1).value = self.data1
        self.sheet.cell(1,2).value = self.data2
        self.sheet.cell(2,1).value = self.data3
        self.sheet.cell(2,2).value = self.data4
        self.sheet.cell(3,1).value = self.data5
        self.sheet.cell(3,2).value = self.data6
        self.sheet.cell(4,1).value = self.data7
        self.sheet.cell(4,2).value = self.data8
        self.sheet.cell(1,4).value = self.data9
        self.sheet.cell(2,4).value = self.data10
        self.sheet.cell(3,4).value = self.data11
        self.sheet.cell(4,4).value = self.data12
        self.op.save(self.excelname)
        self.op.close()

    def read_data(self):
        data=[]
        for row in range(1,self.sheet.max_row+1):
            for column in range(1,self.sheet.max_column+1):
                value = self.sheet.cell(row,column).value
                data.append(value)
        self.op.close()
        return data


if __name__ == '__main__':
    ce = Create_Excel('xmm','pyexcel.xlsx')
    ce.write_data()
    data = ce.read_data()
    print(data)


