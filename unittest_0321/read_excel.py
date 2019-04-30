from openpyxl import load_workbook

class Read_Excel:

    wb = load_workbook("pyexcel.xlsx")
    sheet = wb['xmm']

    def readdata(self):
        list1 = []
        for row in range(1,self.sheet.max_row+1):
            # list2=[]
            # for column in range(1,self.sheet.max_column+1):
            dict={
                'params':self.sheet.cell(row,1).value,
                'expect':self.sheet.cell(row,2).value,
                'code_id':self.sheet.cell(row,4).value
                }
            list1.append(dict)
        self.wb.close()
        return list1

    def writedata(self,row,result):
        self.sheet.cell(row,3).value = result
        self.wb.save("pyexcel.xlsx")
        self.wb.close()


if __name__ == '__main__':
    print(Read_Excel().readdata())